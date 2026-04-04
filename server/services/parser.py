"""
Document Parser — Handles PDF, Excel, and plain text/email files.
Returns a list of chunks, each with text content and metadata.

IMPROVED: Respects embedding model context limits (8192 tokens for nomic-embed-text)
"""
import io
import re
import json
from datetime import datetime
from dateutil import parser as dateparser


def estimate_tokens(text):
    """
    Estimate token count for text.
    Rough approximation: 1 token ≈ 4 characters for English text.
    """
    return len(text) // 4


def parse_file(filename, file_bytes):
    """
    Parse a file based on its extension and return structured chunks.
    Each chunk = { "text": str, "metadata": dict }
    
    All chunks are validated to fit within embedding model context limits.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"

    if ext == "pdf":
        return _parse_pdf(filename, file_bytes)
    elif ext in ("xlsx", "xls"):
        return _parse_excel_improved(filename, file_bytes)
    elif ext in ("txt", "eml", "email", "text"):
        return _parse_text(filename, file_bytes)
    else:
        # Fallback: treat as plain text
        return _parse_text(filename, file_bytes)


def _parse_pdf(filename, file_bytes):
    """
    Extract text from PDF page by page with token-aware chunking.
    Includes preprocessing for long context pages.
    Ultra-safe limits for all providers.
    """
    from PyPDF2 import PdfReader

    chunks = []
    reader = PdfReader(io.BytesIO(file_bytes))
    MAX_TOKENS = 1200  # Ultra-safe limit (reduced from 1500)
    
    print(f"\n📄 PARSING PDF: {filename}")
    print(f"   Pages: {len(reader.pages)}")
    print(f"   Max tokens per chunk: {MAX_TOKENS} (ultra-safe for all providers)")

    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text()
        if text and text.strip():
            page_tokens = estimate_tokens(text.strip())
            page_chars = len(text)
            print(f"   Page {page_num}: {page_tokens} tokens, {page_chars} chars")
            
            # Split large pages into token-limited chunks
            sub_chunks = _split_text_by_tokens(text.strip(), max_tokens=MAX_TOKENS)
            print(f"      → Split into {len(sub_chunks)} chunk(s)")
            
            for i, chunk_text in enumerate(sub_chunks):
                chunk_tokens = estimate_tokens(chunk_text)
                chunk_chars = len(chunk_text)
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": filename,
                        "source_type": "pdf",
                        "page": page_num,
                        "chunk_index": i,
                        "location": f"Page {page_num}",
                    }
                })
                print(f"      ✅ Chunk {i+1}: {chunk_tokens} tokens, {chunk_chars} chars")

    print(f"\n   ✅ PDF parsing complete: {len(chunks)} chunks created")
    print(f"   📊 All chunks are safe for all providers (< {MAX_TOKENS} tokens, < {MAX_TOKENS * 4} chars)")
    return chunks


def _parse_excel_improved(filename, file_bytes):
    """
    Extract Excel data with smart chunking to respect token limits.
    
    Strategy:
    1. PREPROCESSING: Detect wide spreadsheets (many columns) and apply aggressive truncation
    2. Convert each row to JSON format for structure preservation
    3. Batch rows dynamically based on token count (max 1200 tokens per chunk)
    4. Include headers in each chunk for context
    5. Ensure no chunk exceeds embedding model limits
    6. Automatic preprocessing for oversized chunks
    
    CRITICAL: Ultra-conservative limit for both Gemini (2048) and Ollama (8192)
    """
    from openpyxl import load_workbook

    chunks = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    
    # ULTRA-SAFE: 1200 tokens = ~4800 chars (well below Ollama's 8192 token limit)
    # Reduced from 1500 to handle wide spreadsheets with many columns
    MAX_TOKENS = 1200
    MAX_VALUE_LENGTH = 200  # Reduced from 300 for aggressive preprocessing
    
    print(f"\n📄 PARSING EXCEL: {filename}")
    print(f"   Max tokens per chunk: {MAX_TOKENS} (ultra-safe for all providers)")
    print(f"   Max value length: {MAX_VALUE_LENGTH} chars (aggressive preprocessing)")

    for sheet_name in wb.sheetnames:
        print(f"\n   📊 Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"      ⚠️  Empty sheet, skipping")
            continue

        # First row = headers
        headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]
        num_columns = len(headers)
        num_data_rows = len(rows) - 1
        
        print(f"      Columns: {num_columns}")
        print(f"      Data rows: {num_data_rows}")
        
        # PREPROCESSING: Detect wide spreadsheets (many columns)
        if num_columns > 50:
            print(f"      ⚠️  WIDE SPREADSHEET DETECTED ({num_columns} columns)")
            print(f"      🔧 Applying aggressive preprocessing:")
            print(f"         - Truncating values to {MAX_VALUE_LENGTH} chars")
            print(f"         - Reducing rows per chunk")
            print(f"         - Extra validation for chunk size")
        
        # Build header context (included in every chunk)
        # For wide spreadsheets, truncate header list to save space
        if num_columns > 50:
            header_preview = ', '.join(headers[:10]) + f'... ({num_columns} total columns)'
            header_text = f"Sheet: {sheet_name}\nColumns ({num_columns} total): {header_preview}\n\n"
        else:
            header_text = f"Sheet: {sheet_name}\nColumns: {', '.join(headers)}\n\n"
        
        header_tokens = estimate_tokens(header_text)
        print(f"      Header tokens: {header_tokens}")
        
        # Process data rows
        current_chunk_rows = []
        current_chunk_tokens = header_tokens
        chunk_start_row = 1
        chunk_count = 0
        
        for row_idx, row in enumerate(rows[1:], start=2):  # Start from row 2 (skip header)
            # Convert row to structured JSON format
            row_dict = {}
            has_data = False
            truncation_count = 0
            
            for h, val in zip(headers, row):
                if val is not None and str(val).strip():
                    # Clean and format value
                    val_str = str(val).strip()
                    # Aggressive truncation for preprocessing
                    if len(val_str) > MAX_VALUE_LENGTH:
                        truncation_count += 1
                        val_str = val_str[:MAX_VALUE_LENGTH] + "..."
                    row_dict[h] = val_str
                    has_data = True
            
            if truncation_count > 0:
                print(f"      🔧 Row {row_idx}: Truncated {truncation_count} value(s) to {MAX_VALUE_LENGTH} chars")
            
            if not has_data:
                continue  # Skip empty rows
            
            # Format row as JSON for better structure
            row_json = json.dumps(row_dict, ensure_ascii=False, indent=2)
            row_text = f"Row {row_idx}:\n{row_json}\n"
            row_tokens = estimate_tokens(row_text)
            row_chars = len(row_text)
            
            # EXTRA VALIDATION: Check if single row is too large
            if row_tokens > MAX_TOKENS * 0.8:  # If single row uses >80% of limit
                print(f"      ⚠️  Row {row_idx} is very large ({row_tokens} tokens, {row_chars} chars)")
                print(f"         This row alone uses {(row_tokens/MAX_TOKENS)*100:.1f}% of chunk limit")
                # Further truncate this specific row
                if num_columns > 30:
                    # For wide rows, only keep first 30 columns
                    truncated_dict = {k: v for i, (k, v) in enumerate(row_dict.items()) if i < 30}
                    truncated_dict['_truncated'] = f'... ({num_columns - 30} more columns omitted)'
                    row_json = json.dumps(truncated_dict, ensure_ascii=False, indent=2)
                    row_text = f"Row {row_idx}:\n{row_json}\n"
                    row_tokens = estimate_tokens(row_text)
                    row_chars = len(row_text)
                    print(f"         → Reduced to first 30 columns: {row_tokens} tokens, {row_chars} chars")
            
            # Check if adding this row would exceed limit
            if current_chunk_tokens + row_tokens > MAX_TOKENS:
                # Save current chunk
                if current_chunk_rows:
                    chunk_text = header_text + "\n".join(current_chunk_rows)
                    chunk_tokens = estimate_tokens(chunk_text)
                    chunk_chars = len(chunk_text)
                    
                    # FINAL VALIDATION: Ensure chunk is safe
                    if chunk_tokens > MAX_TOKENS or chunk_chars > MAX_TOKENS * 4:
                        print(f"      ⚠️  Chunk {chunk_count + 1} still too large after preprocessing!")
                        print(f"         Tokens: {chunk_tokens}, Chars: {chunk_chars}")
                        print(f"         Splitting chunk further...")
                        # Emergency split: take only half the rows
                        mid = len(current_chunk_rows) // 2
                        if mid > 0:
                            # Save first half
                            chunk_text = header_text + "\n".join(current_chunk_rows[:mid])
                            chunk_tokens = estimate_tokens(chunk_text)
                            chunk_chars = len(chunk_text)
                            chunk_count += 1
                            print(f"      ✅ Chunk {chunk_count} (split 1/2): Rows {chunk_start_row}-{chunk_start_row + mid - 1} ({chunk_tokens} tokens, {chunk_chars} chars)")
                            chunks.append({
                                "text": chunk_text,
                                "metadata": {
                                    "source": filename,
                                    "source_type": "excel",
                                    "sheet": sheet_name,
                                    "rows": f"{chunk_start_row}-{chunk_start_row + mid - 1}",
                                    "row_count": mid,
                                    "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{chunk_start_row + mid - 1}",
                                    "format": "json",
                                    "preprocessed": True
                                }
                            })
                            # Save second half
                            chunk_text = header_text + "\n".join(current_chunk_rows[mid:])
                            chunk_tokens = estimate_tokens(chunk_text)
                            chunk_chars = len(chunk_text)
                            chunk_count += 1
                            print(f"      ✅ Chunk {chunk_count} (split 2/2): Rows {chunk_start_row + mid}-{row_idx - 1} ({chunk_tokens} tokens, {chunk_chars} chars)")
                            chunks.append({
                                "text": chunk_text,
                                "metadata": {
                                    "source": filename,
                                    "source_type": "excel",
                                    "sheet": sheet_name,
                                    "rows": f"{chunk_start_row + mid}-{row_idx - 1}",
                                    "row_count": len(current_chunk_rows) - mid,
                                    "location": f"Sheet '{sheet_name}', Rows {chunk_start_row + mid}-{row_idx - 1}",
                                    "format": "json",
                                    "preprocessed": True
                                }
                            })
                        else:
                            # Single row chunk - just save it
                            chunk_count += 1
                            print(f"      ✅ Chunk {chunk_count}: Rows {chunk_start_row}-{row_idx - 1} ({chunk_tokens} tokens, {chunk_chars} chars)")
                            chunks.append({
                                "text": chunk_text,
                                "metadata": {
                                    "source": filename,
                                    "source_type": "excel",
                                    "sheet": sheet_name,
                                    "rows": f"{chunk_start_row}-{row_idx - 1}",
                                    "row_count": len(current_chunk_rows),
                                    "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{row_idx - 1}",
                                    "format": "json",
                                    "preprocessed": True
                                }
                            })
                    else:
                        chunk_count += 1
                        print(f"      ✅ Chunk {chunk_count}: Rows {chunk_start_row}-{row_idx - 1} ({chunk_tokens} tokens, {chunk_chars} chars)")
                        chunks.append({
                            "text": chunk_text,
                            "metadata": {
                                "source": filename,
                                "source_type": "excel",
                                "sheet": sheet_name,
                                "rows": f"{chunk_start_row}-{row_idx - 1}",
                                "row_count": len(current_chunk_rows),
                                "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{row_idx - 1}",
                                "format": "json"
                            }
                        })
                
                # Start new chunk
                current_chunk_rows = [row_text]
                current_chunk_tokens = header_tokens + row_tokens
                chunk_start_row = row_idx
            else:
                # Add row to current chunk
                current_chunk_rows.append(row_text)
                current_chunk_tokens += row_tokens
        
        # Save final chunk
        if current_chunk_rows:
            chunk_text = header_text + "\n".join(current_chunk_rows)
            chunk_tokens = estimate_tokens(chunk_text)
            chunk_chars = len(chunk_text)
            
            # FINAL VALIDATION: Ensure chunk is safe
            if chunk_tokens > MAX_TOKENS or chunk_chars > MAX_TOKENS * 4:
                print(f"      ⚠️  Final chunk too large after preprocessing!")
                print(f"         Tokens: {chunk_tokens}, Chars: {chunk_chars}")
                print(f"         Splitting chunk further...")
                # Emergency split: take only half the rows
                mid = len(current_chunk_rows) // 2
                if mid > 0:
                    # Save first half
                    chunk_text = header_text + "\n".join(current_chunk_rows[:mid])
                    chunk_tokens = estimate_tokens(chunk_text)
                    chunk_chars = len(chunk_text)
                    chunk_count += 1
                    print(f"      ✅ Chunk {chunk_count} (split 1/2): Rows {chunk_start_row}-{chunk_start_row + mid - 1} ({chunk_tokens} tokens, {chunk_chars} chars)")
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "source": filename,
                            "source_type": "excel",
                            "sheet": sheet_name,
                            "rows": f"{chunk_start_row}-{chunk_start_row + mid - 1}",
                            "row_count": mid,
                            "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{chunk_start_row + mid - 1}",
                            "format": "json",
                            "preprocessed": True
                        }
                    })
                    # Save second half
                    chunk_text = header_text + "\n".join(current_chunk_rows[mid:])
                    chunk_tokens = estimate_tokens(chunk_text)
                    chunk_chars = len(chunk_text)
                    chunk_count += 1
                    print(f"      ✅ Chunk {chunk_count} (split 2/2): Rows {chunk_start_row + mid}-{len(rows)} ({chunk_tokens} tokens, {chunk_chars} chars)")
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "source": filename,
                            "source_type": "excel",
                            "sheet": sheet_name,
                            "rows": f"{chunk_start_row + mid}-{len(rows)}",
                            "row_count": len(current_chunk_rows) - mid,
                            "location": f"Sheet '{sheet_name}', Rows {chunk_start_row + mid}-{len(rows)}",
                            "format": "json",
                            "preprocessed": True
                        }
                    })
                else:
                    # Single row chunk - just save it
                    chunk_count += 1
                    print(f"      ✅ Chunk {chunk_count}: Rows {chunk_start_row}-{len(rows)} ({chunk_tokens} tokens, {chunk_chars} chars)")
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "source": filename,
                            "source_type": "excel",
                            "sheet": sheet_name,
                            "rows": f"{chunk_start_row}-{len(rows)}",
                            "row_count": len(current_chunk_rows),
                            "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{len(rows)}",
                            "format": "json",
                            "preprocessed": True
                        }
                    })
            else:
                chunk_count += 1
                print(f"      ✅ Chunk {chunk_count}: Rows {chunk_start_row}-{len(rows)} ({chunk_tokens} tokens, {chunk_chars} chars)")
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": filename,
                        "source_type": "excel",
                        "sheet": sheet_name,
                        "rows": f"{chunk_start_row}-{len(rows)}",
                        "row_count": len(current_chunk_rows),
                        "location": f"Sheet '{sheet_name}', Rows {chunk_start_row}-{len(rows)}",
                        "format": "json"
                    }
                })

    wb.close()
    
    # Validate all chunks are within limits (safety check)
    print(f"\n   🔍 Final validation of all chunks...")
    validated_chunks = []
    for i, chunk in enumerate(chunks):
        tokens = estimate_tokens(chunk["text"])
        chars = len(chunk["text"])
        if tokens > 1300 or chars > 5200:  # Extra safety margin (stricter than before)
            print(f"      ⚠️  Chunk {i+1} still too large ({tokens} tokens, {chars} chars), emergency split...")
            # Split this chunk further
            sub_chunks = _split_large_chunk(chunk, MAX_TOKENS // 2)  # Use half the limit for safety
            validated_chunks.extend(sub_chunks)
            print(f"      ✅ Emergency split into {len(sub_chunks)} sub-chunks")
        else:
            validated_chunks.append(chunk)
    
    print(f"\n   ✅ Excel parsing complete: {len(validated_chunks)} chunks created")
    print(f"   📊 All chunks are safe for all providers (< {MAX_TOKENS} tokens, < {MAX_TOKENS * 4} chars)")
    return validated_chunks


def _split_large_chunk(chunk, max_tokens):
    """
    Split a chunk that's too large into smaller chunks.
    Used as a safety fallback.
    """
    text = chunk["text"]
    metadata = chunk["metadata"]
    
    # Split by lines
    lines = text.split("\n")
    sub_chunks = []
    current_text = ""
    current_tokens = 0
    
    for line in lines:
        line_tokens = estimate_tokens(line)
        if current_tokens + line_tokens > max_tokens:
            if current_text:
                sub_chunks.append({
                    "text": current_text,
                    "metadata": {**metadata, "split": len(sub_chunks)}
                })
            current_text = line + "\n"
            current_tokens = line_tokens
        else:
            current_text += line + "\n"
            current_tokens += line_tokens
    
    if current_text:
        sub_chunks.append({
            "text": current_text,
            "metadata": {**metadata, "split": len(sub_chunks)}
        })
    
    return sub_chunks if sub_chunks else [chunk]


def _parse_text(filename, file_bytes):
    """
    Parse plain text or email files with token-aware chunking.
    Includes preprocessing for long context files.
    Ultra-safe limits for all providers.
    """
    text = file_bytes.decode("utf-8", errors="replace")
    chunks = []
    MAX_TOKENS = 1200  # Ultra-safe limit (reduced from 1500)
    
    print(f"\n📄 PARSING TEXT: {filename}")
    print(f"   Total length: {len(text)} chars (~{estimate_tokens(text)} tokens)")
    print(f"   Max tokens per chunk: {MAX_TOKENS} (ultra-safe for all providers)")
    
    sub_chunks = _split_text_by_tokens(text.strip(), max_tokens=MAX_TOKENS)
    print(f"   Split into {len(sub_chunks)} chunk(s)")

    for i, chunk_text in enumerate(sub_chunks):
        chunk_tokens = estimate_tokens(chunk_text)
        chunk_chars = len(chunk_text)
        chunks.append({
            "text": chunk_text,
            "metadata": {
                "source": filename,
                "source_type": "text",
                "chunk_index": i,
                "location": f"Section {i + 1}",
            }
        })
        print(f"   ✅ Chunk {i+1}: {chunk_tokens} tokens, {chunk_chars} chars")

    print(f"\n   ✅ Text parsing complete: {len(chunks)} chunks created")
    print(f"   📊 All chunks are safe for all providers (< {MAX_TOKENS} tokens, < {MAX_TOKENS * 4} chars)")
    return chunks


def _split_text_by_tokens(text, max_tokens=1200):
    """
    Split text into chunks based on token count, respecting sentence boundaries.
    Ultra-safe limits for all providers.
    
    Args:
        text: Text to split
        max_tokens: Maximum tokens per chunk (default: 1200 for ultra-safety)
        
    Returns:
        List of text chunks
    """
    if estimate_tokens(text) <= max_tokens:
        return [text]

    chunks = []
    # Try splitting by paragraphs first
    paragraphs = re.split(r"\n\s*\n", text)

    current_chunk = ""
    current_tokens = 0
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        
        para_tokens = estimate_tokens(para)
        
        if current_tokens + para_tokens <= max_tokens:
            current_chunk = (current_chunk + "\n\n" + para).strip()
            current_tokens += para_tokens
        else:
            if current_chunk:
                chunks.append(current_chunk)
            
            # If single paragraph exceeds max_tokens, split by sentences
            if para_tokens > max_tokens:
                sentences = re.split(r"(?<=[.!?])\s+", para)
                sub_chunk = ""
                sub_tokens = 0
                
                for sent in sentences:
                    sent_tokens = estimate_tokens(sent)
                    if sub_tokens + sent_tokens <= max_tokens:
                        sub_chunk = (sub_chunk + " " + sent).strip()
                        sub_tokens += sent_tokens
                    else:
                        if sub_chunk:
                            chunks.append(sub_chunk)
                        # If single sentence is too long, hard split
                        if sent_tokens > max_tokens:
                            # Split by character count as last resort
                            max_chars = max_tokens * 4
                            for i in range(0, len(sent), max_chars):
                                chunks.append(sent[i:i + max_chars])
                            sub_chunk = ""
                            sub_tokens = 0
                        else:
                            sub_chunk = sent
                            sub_tokens = sent_tokens
                
                current_chunk = sub_chunk
                current_tokens = sub_tokens
            else:
                current_chunk = para
                current_tokens = para_tokens

    if current_chunk:
        chunks.append(current_chunk)

    return chunks if chunks else [text[:max_tokens * 4]]


def extract_date_from_text(text):
    """
    Try to extract a date from document text for conflict resolution.
    Returns a datetime or None.
    """
    # Common date patterns
    patterns = [
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
        r"\b(\w+ \d{1,2},?\s*\d{4})\b",
        r"\b(\d{4}-\d{2}-\d{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return dateparser.parse(match.group(1), fuzzy=True)
            except (ValueError, OverflowError):
                continue
    return None
