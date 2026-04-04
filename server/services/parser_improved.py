"""
IMPROVED Document Parser — Uses PyMuPDF4LLM for better PDF extraction.
Handles PDF, Excel, and plain text/email files with structure preservation.
"""
import io
import re
import tempfile
import os
from datetime import datetime
from dateutil import parser as dateparser


def parse_file(filename, file_bytes):
    """
    Parse a file based on its extension and return structured chunks.
    Each chunk = { "text": str, "metadata": dict }
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"

    if ext == "pdf":
        return _parse_pdf_improved(filename, file_bytes)
    elif ext in ("xlsx", "xls"):
        return _parse_excel(filename, file_bytes)
    elif ext in ("txt", "eml", "email", "text"):
        return _parse_text(filename, file_bytes)
    else:
        # Fallback: treat as plain text
        return _parse_text(filename, file_bytes)


def _parse_pdf_improved(filename, file_bytes):
    """
    Extract PDF with structure preservation using PyMuPDF4LLM.
    Preserves headers, tables, lists, and formatting in markdown.
    """
    try:
        import pymupdf4llm
    except ImportError:
        print("⚠️  pymupdf4llm not installed, falling back to PyPDF2")
        return _parse_pdf_fallback(filename, file_bytes)
    
    chunks = []
    
    # PyMuPDF4LLM needs a file path, so create temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    
    try:
        # Extract with structure preservation
        md_text = pymupdf4llm.to_markdown(
            tmp_path,
            page_chunks=True,  # Get per-page chunks with metadata
            write_images=False,  # Skip images for now
            show_progress=False
        )
        
        # md_text is a list of dicts: [{"page": 1, "text": "...", "metadata": {...}}, ...]
        if isinstance(md_text, str):
            # Single string output - split by pages manually
            md_text = [{"page": 1, "text": md_text}]
        
        for page_data in md_text:
            page_num = page_data.get("page", 1)
            text = page_data.get("text", "")
            
            if not text.strip():
                continue
            
            # Extract section headers for better metadata
            sections = _extract_markdown_sections(text)
            
            # Smart chunking by markdown structure
            page_chunks = _chunk_markdown_smart(text, max_tokens=512)
            
            for i, chunk_text in enumerate(page_chunks):
                # Get the most relevant section for this chunk
                chunk_section = _find_chunk_section(chunk_text, sections)
                
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": filename,
                        "source_type": "pdf",
                        "page": page_num,
                        "chunk_index": i,
                        "location": f"Page {page_num}" + (f" - {chunk_section}" if chunk_section else ""),
                        "section": chunk_section,
                        "format": "markdown",  # Flag for special handling
                        "has_table": "| " in chunk_text,  # Detect tables
                        "has_list": any(line.strip().startswith(("-", "*", "1.")) for line in chunk_text.split("\n")),
                    }
                })
    finally:
        # Clean up temp file
        try:
            os.unlink(tmp_path)
        except:
            pass
    
    return chunks


def _parse_pdf_fallback(filename, file_bytes):
    """Fallback PDF parser using PyPDF2 if PyMuPDF4LLM not available."""
    from PyPDF2 import PdfReader

    chunks = []
    reader = PdfReader(io.BytesIO(file_bytes))

    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text()
        if text and text.strip():
            # Split large pages into smaller chunks (~500 chars)
            sub_chunks = _split_text(text.strip(), max_chars=500)
            for i, chunk_text in enumerate(sub_chunks):
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": filename,
                        "source_type": "pdf",
                        "page": page_num,
                        "chunk_index": i,
                        "location": f"Page {page_num}",
                    }
                })

    return chunks


def _extract_markdown_sections(markdown_text):
    """
    Extract section headers from markdown text.
    Returns list of {"level": int, "title": str, "line": int}
    """
    sections = []
    for line_num, line in enumerate(markdown_text.split("\n")):
        if line.startswith("#"):
            level = len(line) - len(line.lstrip("#"))
            title = line.lstrip("#").strip()
            if title:
                sections.append({
                    "level": level,
                    "title": title,
                    "line": line_num
                })
    return sections


def _find_chunk_section(chunk_text, sections):
    """Find the most relevant section header for a chunk."""
    # Look for headers in the chunk itself
    for section in sections:
        if section["title"] in chunk_text:
            return section["title"]
    return None


def _chunk_markdown_smart(text, max_tokens=512, overlap_tokens=50):
    """
    Smart chunking that respects markdown structure.
    - Keeps sections together
    - Preserves tables
    - Adds overlap for context
    - Respects token limits
    """
    # Rough token estimation: 1 token ≈ 4 characters
    max_chars = max_tokens * 4
    overlap_chars = overlap_tokens * 4
    
    # Split by major sections (## headers)
    section_pattern = r"\n(?=#{1,3}\s)"
    sections = re.split(section_pattern, text)
    
    chunks = []
    current_chunk = ""
    
    for section in sections:
        section = section.strip()
        if not section:
            continue
        
        # Check if section has a table - keep tables together
        has_table = "| " in section and "|---" in section
        
        if has_table:
            # Try to keep entire table in one chunk
            table_start = section.find("|")
            table_end = section.rfind("|") + 1
            
            if table_end - table_start < max_chars:
                # Table fits, keep it together
                if current_chunk:
                    chunks.append(current_chunk)
                    current_chunk = ""
                chunks.append(section)
                continue
        
        # Normal section processing
        if len(current_chunk) + len(section) + 2 <= max_chars:
            current_chunk = (current_chunk + "\n\n" + section).strip()
        else:
            if current_chunk:
                chunks.append(current_chunk)
            
            # If section itself is too large, split by paragraphs
            if len(section) > max_chars:
                paras = section.split("\n\n")
                sub_chunk = ""
                for para in paras:
                    if len(sub_chunk) + len(para) + 2 <= max_chars:
                        sub_chunk = (sub_chunk + "\n\n" + para).strip()
                    else:
                        if sub_chunk:
                            chunks.append(sub_chunk)
                        # If single paragraph is too large, hard split
                        if len(para) > max_chars:
                            chunks.append(para[:max_chars])
                            sub_chunk = para[max_chars - overlap_chars:]
                        else:
                            sub_chunk = para
                current_chunk = sub_chunk
            else:
                current_chunk = section
    
    if current_chunk:
        chunks.append(current_chunk)
    
    # Add overlap between chunks for better context
    if len(chunks) > 1:
        chunks = _add_chunk_overlap(chunks, overlap_chars)
    
    return chunks if chunks else [text[:max_chars]]


def _add_chunk_overlap(chunks, overlap_chars):
    """Add overlap between consecutive chunks for better context."""
    overlapped = []
    for i, chunk in enumerate(chunks):
        if i == 0:
            overlapped.append(chunk)
        else:
            # Add end of previous chunk to start of current
            prev_end = chunks[i-1][-overlap_chars:].strip()
            overlapped.append(prev_end + "\n\n" + chunk)
    return overlapped


def _parse_excel(filename, file_bytes):
    """
    Extract Excel data row by row, preserving column headers for context.
    This is critical — tabular data must be readable as natural language.
    """
    from openpyxl import load_workbook

    chunks = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row = headers
        headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(rows[0])]

        # Group rows into chunks of 5 for better context
        batch_size = 5
        for batch_start in range(1, len(rows), batch_size):
            batch_rows = rows[batch_start: batch_start + batch_size]
            lines = []
            for row_idx, row in enumerate(batch_rows, start=batch_start + 1):
                row_parts = []
                for h, val in zip(headers, row):
                    if val is not None:
                        row_parts.append(f"{h}: {val}")
                if row_parts:
                    lines.append(f"Row {row_idx}: " + " | ".join(row_parts))

            if lines:
                chunk_text = f"Sheet: {sheet_name}\n" + "\n".join(lines)
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": filename,
                        "source_type": "excel",
                        "sheet": sheet_name,
                        "rows": f"{batch_start + 1}-{batch_start + len(batch_rows)}",
                        "chunk_index": batch_start // batch_size,
                        "location": f"Sheet '{sheet_name}', Rows {batch_start + 1}-{batch_start + len(batch_rows)}",
                    }
                })

    wb.close()
    return chunks


def _parse_text(filename, file_bytes):
    """Parse plain text or email files."""
    text = file_bytes.decode("utf-8", errors="replace")
    chunks = []
    sub_chunks = _split_text(text.strip(), max_chars=500)

    for i, chunk_text in enumerate(sub_chunks):
        chunks.append({
            "text": chunk_text,
            "metadata": {
                "source": filename,
                "source_type": "text",
                "chunk_index": i,
                "location": f"Section {i + 1}",
            }
        })

    return chunks


def _split_text(text, max_chars=500):
    """Split text into chunks at sentence/paragraph boundaries."""
    if len(text) <= max_chars:
        return [text]

    chunks = []
    # Try splitting by paragraphs first
    paragraphs = re.split(r"\n\s*\n", text)

    current_chunk = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(current_chunk) + len(para) + 2 <= max_chars:
            current_chunk = (current_chunk + "\n\n" + para).strip()
        else:
            if current_chunk:
                chunks.append(current_chunk)
            # If single paragraph exceeds max_chars, split by sentences
            if len(para) > max_chars:
                sentences = re.split(r"(?<=[.!?])\s+", para)
                sub = ""
                for sent in sentences:
                    if len(sub) + len(sent) + 1 <= max_chars:
                        sub = (sub + " " + sent).strip()
                    else:
                        if sub:
                            chunks.append(sub)
                        sub = sent
                current_chunk = sub
            else:
                current_chunk = para

    if current_chunk:
        chunks.append(current_chunk)

    return chunks if chunks else [text[:max_chars]]


def extract_date_from_text(text):
    """
    Try to extract a date from document text for conflict resolution.
    Returns a datetime or None.
    """
    # Common date patterns
    patterns = [
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
        r"\b(\w+ \d{1,2},?\s*\d{4})\b",
        r"\b(\d{4}-\d{2}-\d{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return dateparser.parse(match.group(1), fuzzy=True)
            except (ValueError, OverflowError):
                continue
    return None
