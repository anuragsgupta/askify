"""Excel parsing module for extracting rows as natural language strings."""

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class ExcelRow:
    """Represents a parsed Excel row with metadata."""
    source: str
    sheet_name: str
    row_number: int
    client: Optional[str]
    doc_date: datetime
    doc_type: str
    content: str


def parse_date_from_filename(filename: str) -> Optional[datetime]:
    """
    Extracts date from filename patterns.
    
    Supported patterns:
    - Pricing_2024-03-15.xlsx → 2024-03-15
    - Sheet_March2024.xlsx → 2024-03-01
    - File_20240315.xlsx → 2024-03-15
    
    Args:
        filename: Name of the file (with or without path)
        
    Returns:
        datetime object if date found, None otherwise
    """
    filename = Path(filename).stem  # Remove extension and path
    
    # Pattern 1: YYYY-MM-DD (e.g., 2024-03-15)
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', filename)
    if match:
        year, month, day = match.groups()
        return datetime(int(year), int(month), int(day))
    
    # Pattern 2: YYYYMMDD (e.g., 20240315)
    match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if match:
        year, month, day = match.groups()
        return datetime(int(year), int(month), int(day))
    
    # Pattern 3: Month name + Year (e.g., March2024, Jan2024)
    month_names = {
        'january': 1, 'jan': 1,
        'february': 2, 'feb': 2,
        'march': 3, 'mar': 3,
        'april': 4, 'apr': 4,
        'may': 5,
        'june': 6, 'jun': 6,
        'july': 7, 'jul': 7,
        'august': 8, 'aug': 8,
        'september': 9, 'sep': 9, 'sept': 9,
        'october': 10, 'oct': 10,
        'november': 11, 'nov': 11,
        'december': 12, 'dec': 12
    }
    
    for month_name, month_num in month_names.items():
        pattern = rf'{month_name}[_\s-]?(\d{{4}})'
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            return datetime(year, month_num, 1)
    
    # Pattern 4: Year only (e.g., Pricing_2024.xlsx)
    match = re.search(r'[_\s-](\d{4})[_\s-]', filename)
    if match:
        year = int(match.group(1))
        return datetime(year, 1, 1)
    
    return None


def extract_date_from_column(row_dict: Dict[str, Any]) -> Optional[datetime]:
    """
    Extracts date from "Date" column if present in the row.
    
    Args:
        row_dict: Dictionary mapping column names to values
        
    Returns:
        datetime object if Date column found and valid, None otherwise
    """
    # Look for common date column names (case-insensitive)
    date_column_names = ['date', 'Date', 'DATE', 'effective_date', 'Effective Date']
    
    for col_name in date_column_names:
        if col_name in row_dict:
            value = row_dict[col_name]
            
            # If already a datetime object
            if isinstance(value, datetime):
                return value
            
            # If string, try to parse
            if isinstance(value, str):
                # Try common date formats
                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                    try:
                        return datetime.strptime(value, fmt)
                    except ValueError:
                        continue
    
    return None


def row_to_natural_language(row_dict: Dict[str, Any]) -> str:
    """
    Converts row dict to natural language preserving all column values.
    
    Args:
        row_dict: Dictionary mapping column names to cell values
        
    Returns:
        Natural language string like "Client: Acme Corp, Product: Widget, Price: $500"
    """
    parts = []
    
    for col_name, value in row_dict.items():
        if value is None or value == '':
            continue
        
        # Format value based on type
        if isinstance(value, (int, float)):
            # Check if it looks like currency (column name contains 'price', 'cost', etc.)
            if any(keyword in col_name.lower() for keyword in ['price', 'cost', 'fee', 'amount']):
                formatted_value = f"${value:,.2f}" if isinstance(value, float) else f"${value:,}"
            else:
                formatted_value = str(value)
        elif isinstance(value, datetime):
            formatted_value = value.strftime('%Y-%m-%d')
        else:
            formatted_value = str(value)
        
        parts.append(f"{col_name}: {formatted_value}")
    
    return ', '.join(parts)


def extract_excel_rows(file_path: str) -> List[ExcelRow]:
    """
    Converts each Excel data row into natural language string.
    Uses openpyxl with data_only=True to get computed formula values.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        List of ExcelRow objects with metadata:
        - source: filename
        - sheet_name: worksheet name
        - row_number: 1-indexed row number
        - client: extracted from "Client" column if present
        - doc_date: from "Date" column or file metadata
        - doc_type: "excel"
        - content: natural language string like "Client: Acme Corp, Product: Widget, Price: $500"
    """
    # Load workbook with data_only=True to get computed formula values
    wb = load_workbook(file_path, data_only=True)
    filename = Path(file_path).name
    
    # Try to extract date from filename first
    doc_date_fallback = parse_date_from_filename(filename)
    if not doc_date_fallback:
        # Fallback to file modification time
        doc_date_fallback = datetime.fromtimestamp(Path(file_path).stat().st_mtime)
    
    excel_rows = []
    
    # Process each worksheet
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        
        # Skip empty sheets
        if sheet.max_row < 2:  # Need at least header + 1 data row
            continue
        
        # Extract header row (assume first row is header)
        headers = []
        for cell in sheet[1]:
            header_value = cell.value
            if header_value is None:
                headers.append(f"Column_{cell.column}")
            else:
                headers.append(str(header_value).strip())
        
        # Process data rows (starting from row 2)
        for row_idx in range(2, sheet.max_row + 1):
            row_cells = sheet[row_idx]
            
            # Build dictionary mapping column names to values
            row_dict = {}
            for col_idx, cell in enumerate(row_cells):
                if col_idx < len(headers):
                    row_dict[headers[col_idx]] = cell.value
            
            # Skip completely empty rows
            if all(v is None or v == '' for v in row_dict.values()):
                continue
            
            # Extract client from "Client" column if present
            client = None
            for col_name in ['Client', 'client', 'CLIENT', 'Customer', 'customer']:
                if col_name in row_dict and row_dict[col_name]:
                    client = str(row_dict[col_name])
                    break
            
            # Extract date from "Date" column or use fallback
            doc_date = extract_date_from_column(row_dict)
            if not doc_date:
                doc_date = doc_date_fallback
            
            # Convert row to natural language
            content = row_to_natural_language(row_dict)
            
            # Create ExcelRow object
            excel_rows.append(ExcelRow(
                source=filename,
                sheet_name=sheet_name,
                row_number=row_idx,
                client=client,
                doc_date=doc_date,
                doc_type="excel",
                content=content
            ))
    
    wb.close()
    return excel_rows
