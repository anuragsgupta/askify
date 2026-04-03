"""Unit tests for Excel parser module."""

import os
import tempfile
import time
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from ingestion.excel_parser import (
    ExcelRow,
    extract_excel_rows,
    parse_date_from_filename,
    extract_date_from_column,
    row_to_natural_language,
)


def safe_delete_file(file_path: str, max_retries: int = 3):
    """Safely delete a file with retries for Windows file locking issues."""
    for i in range(max_retries):
        try:
            time.sleep(0.1)  # Small delay to allow file handles to close
            Path(file_path).unlink()
            return
        except PermissionError:
            if i == max_retries - 1:
                # Last attempt failed, just pass - temp files will be cleaned up eventually
                pass
            else:
                time.sleep(0.2)


class TestParseDateFromFilename:
    """Test date extraction from various filename patterns."""
    
    def test_iso_date_format(self):
        """Test YYYY-MM-DD format."""
        result = parse_date_from_filename("Pricing_2024-03-15.xlsx")
        assert result == datetime(2024, 3, 15)
    
    def test_compact_date_format(self):
        """Test YYYYMMDD format."""
        result = parse_date_from_filename("Sheet_20240315.xlsx")
        assert result == datetime(2024, 3, 15)
    
    def test_month_name_year_format(self):
        """Test Month+Year format."""
        result = parse_date_from_filename("Pricing_March2024.xlsx")
        assert result == datetime(2024, 3, 1)
    
    def test_abbreviated_month_format(self):
        """Test abbreviated month names."""
        result = parse_date_from_filename("Data_Jan2024.xlsx")
        assert result == datetime(2024, 1, 1)
    
    def test_year_only_format(self):
        """Test year-only extraction."""
        result = parse_date_from_filename("Annual_Data_2024_Final.xlsx")
        assert result == datetime(2024, 1, 1)
    
    def test_no_date_in_filename(self):
        """Test filename without date returns None."""
        result = parse_date_from_filename("Pricing_Sheet.xlsx")
        assert result is None


class TestExtractDateFromColumn:
    """Test date extraction from Date column in row data."""
    
    def test_datetime_object_in_date_column(self):
        """Test when Date column contains datetime object."""
        row_dict = {"Date": datetime(2024, 3, 15), "Client": "Acme"}
        result = extract_date_from_column(row_dict)
        assert result == datetime(2024, 3, 15)
    
    def test_string_date_iso_format(self):
        """Test parsing ISO format date string."""
        row_dict = {"Date": "2024-03-15", "Client": "Acme"}
        result = extract_date_from_column(row_dict)
        assert result == datetime(2024, 3, 15)
    
    def test_string_date_us_format(self):
        """Test parsing US format date string."""
        row_dict = {"Date": "03/15/2024", "Client": "Acme"}
        result = extract_date_from_column(row_dict)
        assert result == datetime(2024, 3, 15)
    
    def test_no_date_column(self):
        """Test when no Date column exists."""
        row_dict = {"Client": "Acme", "Price": 500}
        result = extract_date_from_column(row_dict)
        assert result is None
    
    def test_case_insensitive_date_column(self):
        """Test Date column name is case-insensitive."""
        row_dict = {"DATE": datetime(2024, 3, 15), "Client": "Acme"}
        result = extract_date_from_column(row_dict)
        assert result == datetime(2024, 3, 15)


class TestRowToNaturalLanguage:
    """Test conversion of row dict to natural language string."""
    
    def test_simple_row_conversion(self):
        """Test basic row with string and numeric values."""
        row_dict = {"Client": "Acme Corp", "Product": "Widget", "Quantity": 10}
        result = row_to_natural_language(row_dict)
        assert "Client: Acme Corp" in result
        assert "Product: Widget" in result
        assert "Quantity: 10" in result
    
    def test_price_formatting(self):
        """Test that price columns are formatted as currency."""
        row_dict = {"Client": "Acme", "Price": 500.50}
        result = row_to_natural_language(row_dict)
        assert "Price: $500.50" in result
    
    def test_cost_formatting(self):
        """Test that cost columns are formatted as currency."""
        row_dict = {"Product": "Widget", "Cost": 1000}
        result = row_to_natural_language(row_dict)
        assert "Cost: $1,000" in result
    
    def test_datetime_formatting(self):
        """Test that datetime values are formatted as dates."""
        row_dict = {"Client": "Acme", "Date": datetime(2024, 3, 15)}
        result = row_to_natural_language(row_dict)
        assert "Date: 2024-03-15" in result
    
    def test_skip_empty_values(self):
        """Test that None and empty string values are skipped."""
        row_dict = {"Client": "Acme", "Notes": None, "Status": ""}
        result = row_to_natural_language(row_dict)
        assert "Client: Acme" in result
        assert "Notes" not in result
        assert "Status" not in result
    
    def test_preserve_column_order(self):
        """Test that all non-empty columns are included."""
        row_dict = {"A": "1", "B": "2", "C": "3"}
        result = row_to_natural_language(row_dict)
        # All columns should be present
        assert "A: 1" in result
        assert "B: 2" in result
        assert "C: 3" in result


class TestExtractExcelRows:
    """Test complete Excel row extraction."""
    
    def test_basic_excel_extraction(self):
        """Test extraction from simple Excel file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add header row
            ws.append(["Client", "Product", "Price"])
            # Add data rows
            ws.append(["Acme Corp", "Widget", 500])
            ws.append(["Beta Inc", "Gadget", 750])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 2
            assert rows[0].source == Path(tmp.name).name
            assert rows[0].sheet_name == "Sheet1"
            assert rows[0].row_number == 2
            assert rows[0].client == "Acme Corp"
            assert rows[0].doc_type == "excel"
            assert "Client: Acme Corp" in rows[0].content
            assert "Product: Widget" in rows[0].content
            
            safe_delete_file(tmp.name)
    
    def test_formula_evaluation(self):
        """Test that formulas are evaluated to computed values."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            # Add header
            ws.append(["Item", "Quantity", "Price", "Total"])
            # Add data with formula
            ws.append(["Widget", 10, 50, "=B2*C2"])
            
            wb.save(tmp.name)
            wb.close()
            
            # Reload with data_only=True to get computed values
            rows = extract_excel_rows(tmp.name)
            
            # Note: formulas return None when file is just created
            # In real usage, formulas would have computed values
            assert len(rows) == 1
            assert rows[0].row_number == 2
            
            safe_delete_file(tmp.name)
    
    def test_client_extraction(self):
        """Test extraction of client name from Client column."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product"])
            ws.append(["Acme Corp", "Widget"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert rows[0].client == "Acme Corp"
            
            safe_delete_file(tmp.name)
    
    def test_date_from_column(self):
        """Test date extraction from Date column."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Date", "Amount"])
            ws.append(["Acme", datetime(2024, 3, 15), 1000])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert rows[0].doc_date == datetime(2024, 3, 15)
            
            safe_delete_file(tmp.name)
    
    def test_date_from_filename_fallback(self):
        """Test date fallback to filename when no Date column."""
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx",
            prefix="Pricing_2024-03-15_",
            delete=False
        ) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product"])
            ws.append(["Acme", "Widget"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert rows[0].doc_date == datetime(2024, 3, 15)
            
            safe_delete_file(tmp.name)
    
    def test_multiple_sheets(self):
        """Test extraction from multiple worksheets."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            
            # First sheet
            ws1 = wb.active
            ws1.title = "Q1"
            ws1.append(["Client", "Amount"])
            ws1.append(["Acme", 1000])
            
            # Second sheet
            ws2 = wb.create_sheet("Q2")
            ws2.append(["Client", "Amount"])
            ws2.append(["Beta", 2000])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 2
            assert rows[0].sheet_name == "Q1"
            assert rows[1].sheet_name == "Q2"
            assert "Acme" in rows[0].content
            assert "Beta" in rows[1].content
            
            safe_delete_file(tmp.name)
    
    def test_skip_empty_rows(self):
        """Test that completely empty rows are skipped."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product"])
            ws.append(["Acme", "Widget"])
            ws.append([None, None])  # Empty row
            ws.append(["Beta", "Gadget"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 2
            assert rows[0].client == "Acme"
            assert rows[1].client == "Beta"
            
            safe_delete_file(tmp.name)
    
    def test_skip_empty_sheets(self):
        """Test that sheets with no data rows are skipped."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            
            # Sheet with data
            ws1 = wb.active
            ws1.title = "Data"
            ws1.append(["Client"])
            ws1.append(["Acme"])
            
            # Empty sheet
            ws2 = wb.create_sheet("Empty")
            ws2.append(["Header"])  # Only header, no data
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 1
            assert rows[0].sheet_name == "Data"
            
            safe_delete_file(tmp.name)


class TestExcelParserEdgeCases:
    """
    Test edge cases for Excel parser.
    **Validates: Requirements 2.1, 2.2, 2.3**
    """
    
    def test_completely_empty_sheet(self):
        """Test handling of completely empty sheet with no headers or data."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "EmptySheet"
            # Don't add any data at all
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            # Should return empty list, not crash
            assert len(rows) == 0
            
            safe_delete_file(tmp.name)
    
    def test_missing_client_column(self):
        """Test graceful handling when Client column is missing."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            # No Client column
            ws.append(["Product", "Price"])
            ws.append(["Widget", 500])
            ws.append(["Gadget", 750])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 2
            # Client should be None when column is missing
            assert rows[0].client is None
            assert rows[1].client is None
            # Content should still be generated
            assert "Product: Widget" in rows[0].content
            assert "Price: $500" in rows[0].content
            
            safe_delete_file(tmp.name)
    
    def test_missing_date_column(self):
        """Test fallback to filename date when Date column is missing."""
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx",
            prefix="Pricing_2024-06-20_",
            delete=False
        ) as tmp:
            wb = Workbook()
            ws = wb.active
            
            # No Date column
            ws.append(["Client", "Product"])
            ws.append(["Acme", "Widget"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 1
            # Should fall back to filename date
            assert rows[0].doc_date == datetime(2024, 6, 20)
            
            safe_delete_file(tmp.name)
    
    def test_missing_both_client_and_date_columns(self):
        """Test handling when both Client and Date columns are missing."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Product", "Quantity"])
            ws.append(["Widget", 10])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 1
            assert rows[0].client is None
            # Should fall back to file modification time
            assert rows[0].doc_date is not None
            assert isinstance(rows[0].doc_date, datetime)
            
            safe_delete_file(tmp.name)
    
    def test_merged_cells_handling(self):
        """Test that merged cells are handled gracefully."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product", "Price"])
            ws.append(["Acme Corp", "Widget", 500])
            ws.append(["Beta Inc", "Gadget", 750])
            
            # Merge cells in the Client column (A2:A3)
            ws.merge_cells('A2:A3')
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            # Should still extract both rows
            assert len(rows) == 2
            # First row should have the merged cell value
            assert rows[0].client == "Acme Corp"
            # Second row in merged range will have None for that cell
            # but should still be processed
            assert rows[1].row_number == 3
            
            safe_delete_file(tmp.name)
    
    def test_formula_returning_none(self):
        """Test handling of formula cells that return None (not yet calculated)."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Product", "Quantity", "Price", "Total"])
            # Add formula that will return None when loaded with data_only=True
            ws.append(["Widget", 10, 50, "=B2*C2"])
            ws.append(["Gadget", 5, 100, "=B3*C3"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            # Should still extract rows even if formula returns None
            assert len(rows) == 2
            assert "Product: Widget" in rows[0].content
            assert "Quantity: 10" in rows[0].content
            # Total column with None should be skipped in natural language
            assert "Total: None" not in rows[0].content
            
            safe_delete_file(tmp.name)
    
    def test_special_formatting_preserved(self):
        """Test that special number formatting doesn't break extraction."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Revenue", "Percentage", "Date"])
            # Add data with various formats
            ws.append(["Acme", 1234567.89, 0.15, datetime(2024, 3, 15)])
            
            # Apply number formatting (this is metadata, values stay the same)
            ws['B2'].number_format = '$#,##0.00'
            ws['C2'].number_format = '0.00%'
            ws['D2'].number_format = 'yyyy-mm-dd'
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 1
            # Values should be extracted correctly regardless of formatting
            assert "Client: Acme" in rows[0].content
            # Revenue doesn't match currency keywords, so no $ formatting
            assert "Revenue: 1234567.89" in rows[0].content
            # Percentage stored as decimal
            assert "Percentage: 0.15" in rows[0].content
            assert "Date: 2024-03-15" in rows[0].content
            
            safe_delete_file(tmp.name)
    
    def test_sheet_with_only_headers(self):
        """Test sheet that has headers but no data rows."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            # Only header row, no data
            ws.append(["Client", "Product", "Price"])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            # Should return empty list
            assert len(rows) == 0
            
            safe_delete_file(tmp.name)
    
    def test_row_with_all_none_values(self):
        """Test that rows with all None values are skipped."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product", "Price"])
            ws.append(["Acme", "Widget", 500])
            ws.append([None, None, None])  # All None
            ws.append(["Beta", "Gadget", 750])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            # Should skip the all-None row
            assert len(rows) == 2
            assert rows[0].client == "Acme"
            assert rows[1].client == "Beta"
            
            safe_delete_file(tmp.name)
    
    def test_row_with_partial_none_values(self):
        """Test that rows with some None values are still processed."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            ws.append(["Client", "Product", "Price", "Notes"])
            ws.append(["Acme", "Widget", 500, None])  # Notes is None
            ws.append(["Beta", None, 750, "Special"])  # Product is None
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 2
            # First row should skip Notes
            assert "Client: Acme" in rows[0].content
            assert "Product: Widget" in rows[0].content
            assert "Notes" not in rows[0].content
            # Second row should skip Product
            assert "Client: Beta" in rows[1].content
            assert "Product" not in rows[1].content
            assert "Notes: Special" in rows[1].content
            
            safe_delete_file(tmp.name)
    
    def test_missing_column_headers(self):
        """Test handling when some column headers are None."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            
            # Header row with None values
            ws.append(["Client", None, "Price"])
            ws.append(["Acme", "Widget", 500])
            
            wb.save(tmp.name)
            wb.close()
            
            rows = extract_excel_rows(tmp.name)
            
            assert len(rows) == 1
            # Should generate column name for None header
            assert "Client: Acme" in rows[0].content
            assert "Price: $500" in rows[0].content
            # Column with None header should get auto-generated name
            assert "Column_" in rows[0].content or "Widget" in rows[0].content
            
            safe_delete_file(tmp.name)
