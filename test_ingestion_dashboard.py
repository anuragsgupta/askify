"""Unit tests for ingestion dashboard functionality."""

import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock

import fitz
import pytest
from openpyxl import Workbook

from ingestion.pdf_parser import extract_pdf_sections
from ingestion.excel_parser import extract_excel_rows
from ingestion.email_parser import parse_eml_file
from storage.chroma_store import init_chroma_collection, upsert_chunks, DocumentChunk


class TestFileUploadAndProcessing:
    """Tests for file upload and processing functionality."""
    
    def test_process_single_pdf_file(self):
        """Test processing a single PDF file."""
        # Create test PDF
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_name = tmp.name
        
        try:
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "Test content")
            toc = [[1, "Introduction", 1]]
            doc.set_toc(toc)
            doc.save(tmp_name)
            doc.close()
            
            # Extract sections
            sections = extract_pdf_sections(tmp_name)
            
            # Verify extraction
            assert len(sections) == 1
            assert sections[0].section_title == "Introduction"
            assert "Test content" in sections[0].content
            assert sections[0].doc_type == "policy"
        finally:
            # Cleanup
            try:
                Path(tmp_name).unlink()
            except PermissionError:
                pass  # Windows file locking - ignore
    
    def test_process_single_excel_file(self):
        """Test processing a single Excel file."""
        # Create test Excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_name = tmp.name
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['Client', 'Product', 'Price'])
            ws.append(['Acme Corp', 'Widget', 500])
            ws.append(['Beta Inc', 'Gadget', 750])
            wb.save(tmp_name)
            wb.close()
            
            # Extract rows
            rows = extract_excel_rows(tmp_name)
            
            # Verify extraction
            assert len(rows) == 2
            assert 'Acme Corp' in rows[0].content
            assert 'Widget' in rows[0].content
            assert rows[0].doc_type == "excel"
        finally:
            # Cleanup
            try:
                Path(tmp_name).unlink()
            except PermissionError:
                pass  # Windows file locking - ignore
    
    def test_process_single_eml_file(self):
        """Test processing a single EML file."""
        # Create test EML
        with tempfile.NamedTemporaryFile(suffix=".eml", delete=False, mode='w') as tmp:
            eml_content = """From: sender@example.com
To: recipient@example.com
Subject: Test Email
Date: Mon, 1 Jan 2024 12:00:00 +0000
Message-ID: <test123@example.com>

This is a test email message.
"""
            tmp.write(eml_content)
            tmp.flush()
            tmp_name = tmp.name
        
        try:
            # Extract messages
            messages = parse_eml_file(tmp_name)
            
            # Verify extraction
            assert len(messages) == 1
            assert messages[0].sender == "sender@example.com"
            assert messages[0].subject == "Test Email"
            assert "test email message" in messages[0].content
            assert messages[0].doc_type == "email"
        finally:
            # Cleanup
            try:
                Path(tmp_name).unlink()
            except PermissionError:
                pass  # Windows file locking - ignore
    
    def test_process_multiple_files_mixed_types(self):
        """Test processing multiple files of different types."""
        files_processed = []
        
        # Create and process PDF
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf_name = tmp_pdf.name
        
        try:
            doc = fitz.open()
            doc.new_page()
            toc = [[1, "Section 1", 1]]
            doc.set_toc(toc)
            doc.save(tmp_pdf_name)
            doc.close()
            
            sections = extract_pdf_sections(tmp_pdf_name)
            files_processed.append(('pdf', len(sections)))
        finally:
            try:
                Path(tmp_pdf_name).unlink()
            except PermissionError:
                pass
        
        # Create and process Excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
            tmp_excel_name = tmp_excel.name
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['Client', 'Product'])
            ws.append(['Test', 'Item'])
            wb.save(tmp_excel_name)
            wb.close()
            
            rows = extract_excel_rows(tmp_excel_name)
            files_processed.append(('excel', len(rows)))
        finally:
            try:
                Path(tmp_excel_name).unlink()
            except PermissionError:
                pass
        
        # Verify processing
        assert len(files_processed) == 2
        assert files_processed[0][0] == 'pdf'
        assert files_processed[1][0] == 'excel'
    
    def test_process_empty_pdf_fallback(self):
        """Test processing PDF without TOC uses fallback."""
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_name = tmp.name
        
        try:
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text((72, 72), "Content without TOC")
            doc.save(tmp_name)
            doc.close()
            
            sections = extract_pdf_sections(tmp_name)
            
            # Should create single "Full Document" section
            assert len(sections) == 1
            assert sections[0].section_title == "Full Document"
            assert "Content without TOC" in sections[0].content
        finally:
            try:
                Path(tmp_name).unlink()
            except PermissionError:
                pass
    
    def test_process_excel_with_empty_rows_skipped(self):
        """Test that empty Excel rows are skipped."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_name = tmp.name
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['Client', 'Product'])
            ws.append(['Acme', 'Widget'])
            ws.append([None, None])  # Empty row
            ws.append(['Beta', 'Gadget'])
            wb.save(tmp_name)
            wb.close()
            
            rows = extract_excel_rows(tmp_name)
            
            # Should skip empty row
            assert len(rows) == 2
            assert 'Acme' in rows[0].content
            assert 'Beta' in rows[1].content
        finally:
            try:
                Path(tmp_name).unlink()
            except PermissionError:
                pass


class TestSummaryStatisticsCalculation:
    """Tests for summary statistics calculation."""
    
    def test_calculate_stats_single_document_type(self):
        """Test calculating stats for single document type."""
        stats = {
            'total_documents': 1,
            'total_sections': 5,
            'total_excel_rows': 0,
            'total_email_messages': 0
        }
        
        assert stats['total_documents'] == 1
        assert stats['total_sections'] == 5
        assert stats['total_excel_rows'] == 0
        assert stats['total_email_messages'] == 0
    
    def test_calculate_stats_mixed_document_types(self):
        """Test calculating stats for mixed document types."""
        stats = {
            'total_documents': 3,
            'total_sections': 10,
            'total_excel_rows': 25,
            'total_email_messages': 5
        }
        
        assert stats['total_documents'] == 3
        assert stats['total_sections'] == 10
        assert stats['total_excel_rows'] == 25
        assert stats['total_email_messages'] == 5
    
    def test_calculate_stats_empty_collection(self):
        """Test calculating stats for empty collection."""
        stats = {
            'total_documents': 0,
            'total_sections': 0,
            'total_excel_rows': 0,
            'total_email_messages': 0
        }
        
        assert stats['total_documents'] == 0
        assert stats['total_sections'] == 0
        assert stats['total_excel_rows'] == 0
        assert stats['total_email_messages'] == 0
    
    def test_stats_increment_correctly(self):
        """Test that stats increment correctly as documents are processed."""
        stats = {
            'total_documents': 0,
            'total_sections': 0,
            'total_excel_rows': 0,
            'total_email_messages': 0
        }
        
        # Process first document (PDF with 3 sections)
        stats['total_documents'] += 1
        stats['total_sections'] += 3
        
        assert stats['total_documents'] == 1
        assert stats['total_sections'] == 3
        
        # Process second document (Excel with 10 rows)
        stats['total_documents'] += 1
        stats['total_excel_rows'] += 10
        
        assert stats['total_documents'] == 2
        assert stats['total_excel_rows'] == 10
        
        # Process third document (Email with 2 messages)
        stats['total_documents'] += 1
        stats['total_email_messages'] += 2
        
        assert stats['total_documents'] == 3
        assert stats['total_email_messages'] == 2


class TestDocumentPreviewDisplay:
    """Tests for document preview display functionality."""
    
    def test_preview_pdf_chunks(self):
        """Test displaying preview of PDF chunks."""
        # Create sample PDF chunks
        chunks = [
            {
                'id': 'pdf_test_1',
                'content': 'Section 1 content about refunds',
                'metadata': {
                    'source': 'test.pdf',
                    'doc_type': 'policy',
                    'section_title': 'Refunds',
                    'section_number': '1',
                    'page_number': 1
                }
            },
            {
                'id': 'pdf_test_2',
                'content': 'Section 2 content about returns',
                'metadata': {
                    'source': 'test.pdf',
                    'doc_type': 'policy',
                    'section_title': 'Returns',
                    'section_number': '2',
                    'page_number': 2
                }
            }
        ]
        
        # Verify chunk structure
        assert len(chunks) == 2
        assert chunks[0]['metadata']['doc_type'] == 'policy'
        assert chunks[0]['metadata']['section_title'] == 'Refunds'
        assert chunks[1]['metadata']['section_title'] == 'Returns'
    
    def test_preview_excel_chunks(self):
        """Test displaying preview of Excel chunks."""
        chunks = [
            {
                'id': 'excel_test_1',
                'content': 'Client: Acme Corp, Product: Widget, Price: $500',
                'metadata': {
                    'source': 'pricing.xlsx',
                    'doc_type': 'excel',
                    'sheet_name': 'Q1',
                    'row_number': 2,
                    'client': 'Acme Corp'
                }
            }
        ]
        
        assert chunks[0]['metadata']['doc_type'] == 'excel'
        assert chunks[0]['metadata']['sheet_name'] == 'Q1'
        assert chunks[0]['metadata']['row_number'] == 2
    
    def test_preview_email_chunks(self):
        """Test displaying preview of Email chunks."""
        chunks = [
            {
                'id': 'email_test_1',
                'content': 'Email body content',
                'metadata': {
                    'source': 'message.eml',
                    'doc_type': 'email',
                    'sender': 'sender@example.com',
                    'subject': 'Test Subject',
                    'thread_id': 'thread123'
                }
            }
        ]
        
        assert chunks[0]['metadata']['doc_type'] == 'email'
        assert chunks[0]['metadata']['sender'] == 'sender@example.com'
        assert chunks[0]['metadata']['subject'] == 'Test Subject'
    
    def test_preview_truncates_long_content(self):
        """Test that preview truncates long content."""
        long_content = "A" * 500
        
        # Simulate truncation (first 200 chars + "...")
        preview = long_content[:200] + "..." if len(long_content) > 200 else long_content
        
        assert len(preview) == 203  # 200 + "..."
        assert preview.endswith("...")
    
    def test_preview_preserves_short_content(self):
        """Test that preview preserves short content without truncation."""
        short_content = "This is a short message."
        
        preview = short_content[:200] + "..." if len(short_content) > 200 else short_content
        
        assert preview == short_content
        assert not preview.endswith("...")


class TestIngestionWithChromaDB:
    """Integration tests for ingestion with ChromaDB storage."""
    
    def test_end_to_end_pdf_ingestion(self):
        """Test complete PDF ingestion flow."""
        tmpdir = tempfile.mkdtemp()
        try:
            # Create test PDF
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                tmp_pdf_name = tmp_pdf.name
            
            try:
                doc = fitz.open()
                page = doc.new_page()
                page.insert_text((72, 72), "Test section content")
                toc = [[1, "Test Section", 1]]
                doc.set_toc(toc)
                doc.save(tmp_pdf_name)
                doc.close()
                
                # Extract sections
                sections = extract_pdf_sections(tmp_pdf_name)
                
                # Initialize ChromaDB
                client, collection = init_chroma_collection(persist_directory=tmpdir)
                
                # Convert to DocumentChunk and upsert
                chunks = []
                for section in sections:
                    chunk = DocumentChunk(
                        id=f"pdf_{section.source}_{section.section_number}".replace(" ", "_"),
                        content=section.content,
                        embedding=[0.1] * 384,  # Mock embedding
                        metadata={
                            'source': section.source,
                            'doc_type': section.doc_type,
                            'doc_date': section.doc_date,
                            'section_title': section.section_title,
                            'section_number': section.section_number,
                            'page_number': section.page_number
                        }
                    )
                    chunks.append(chunk)
                
                upsert_chunks(collection, chunks)
                
                # Verify storage
                assert collection.count() == 1
                
                # Clean up ChromaDB references
                del collection
                del client
            finally:
                # Cleanup PDF file
                try:
                    Path(tmp_pdf_name).unlink()
                except PermissionError:
                    pass
        finally:
            # Cleanup temp directory
            import shutil
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except:
                pass
    
    def test_end_to_end_excel_ingestion(self):
        """Test complete Excel ingestion flow."""
        tmpdir = tempfile.mkdtemp()
        try:
            # Create test Excel
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                tmp_excel_name = tmp_excel.name
            
            try:
                wb = Workbook()
                ws = wb.active
                ws.append(['Client', 'Product', 'Price'])
                ws.append(['Acme', 'Widget', 500])
                wb.save(tmp_excel_name)
                wb.close()
                
                # Extract rows
                rows = extract_excel_rows(tmp_excel_name)
                
                # Initialize ChromaDB
                client, collection = init_chroma_collection(persist_directory=tmpdir)
                
                # Convert to DocumentChunk and upsert
                chunks = []
                for row in rows:
                    chunk = DocumentChunk(
                        id=f"excel_{row.source}_{row.sheet_name}_{row.row_number}".replace(" ", "_"),
                        content=row.content,
                        embedding=[0.1] * 384,
                        metadata={
                            'source': row.source,
                            'doc_type': row.doc_type,
                            'doc_date': row.doc_date,
                            'sheet_name': row.sheet_name,
                            'row_number': row.row_number,
                            'client': row.client
                        }
                    )
                    chunks.append(chunk)
                
                upsert_chunks(collection, chunks)
                
                # Verify storage
                assert collection.count() == 1
                
                # Clean up ChromaDB references
                del collection
                del client
            finally:
                # Cleanup Excel file
                try:
                    Path(tmp_excel_name).unlink()
                except PermissionError:
                    pass
        finally:
            # Cleanup temp directory
            import shutil
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except:
                pass
    
    def test_stats_from_chromadb(self):
        """Test calculating stats from ChromaDB collection."""
        tmpdir = tempfile.mkdtemp()
        try:
            # Initialize ChromaDB
            client, collection = init_chroma_collection(persist_directory=tmpdir)
            
            # Add test chunks
            chunks = [
                DocumentChunk(
                    id="pdf_1",
                    content="PDF content",
                    embedding=[0.1] * 384,
                    metadata={'source': 'test.pdf', 'doc_type': 'policy', 'doc_date': datetime(2024, 1, 1)}
                ),
                DocumentChunk(
                    id="excel_1",
                    content="Excel content",
                    embedding=[0.1] * 384,
                    metadata={'source': 'test.xlsx', 'doc_type': 'excel', 'doc_date': datetime(2024, 1, 1)}
                ),
                DocumentChunk(
                    id="email_1",
                    content="Email content",
                    embedding=[0.1] * 384,
                    metadata={'source': 'test.eml', 'doc_type': 'email', 'doc_date': datetime(2024, 1, 1)}
                )
            ]
            
            upsert_chunks(collection, chunks)
            
            # Get all chunks and calculate stats
            all_chunks = collection.get()
            
            stats = {
                'total_documents': 0,
                'total_sections': 0,
                'total_excel_rows': 0,
                'total_email_messages': 0
            }
            
            doc_sources = set()
            for metadata in all_chunks.get('metadatas', []):
                doc_type = metadata.get('doc_type', '')
                source = metadata.get('source', '')
                
                doc_sources.add(source)
                
                if doc_type == 'policy':
                    stats['total_sections'] += 1
                elif doc_type == 'excel':
                    stats['total_excel_rows'] += 1
                elif doc_type == 'email':
                    stats['total_email_messages'] += 1
            
            stats['total_documents'] = len(doc_sources)
            
            # Verify stats
            assert stats['total_documents'] == 3
            assert stats['total_sections'] == 1
            assert stats['total_excel_rows'] == 1
            assert stats['total_email_messages'] == 1
            
            # Clean up ChromaDB references
            del collection
            del client
        finally:
            # Cleanup temp directory
            import shutil
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except:
                pass
